from rdflib import ConjunctiveGraph
from openpyxl import load_workbook
import requests
import gspread


def __get_class_name(temp_uri):
    return temp_uri.replace("http://schema.org/","")


def __add_property(props_dic, prop_desc):
    sdo_uri="http://schema.org/"
    if prop_desc['prop_name'] in props_dic:
        t_prop_name = prop_desc['prop_name']
        props_dic[t_prop_name]['exp_type'].append(prop_desc['exp_type'].replace(sdo_uri,""))
    else:
        props_dic[prop_desc['prop_name']]=prop_desc
        props_dic[prop_desc['prop_name']]['exp_type'] = [prop_desc['exp_type'].replace(sdo_uri,"")]
    return props_dic


def __get_class_props(class_name, graph):
    print("Quering properties of  %s in Schema.org" % class_name)

    qres = graph.query("""prefix schema: <http://schema.org/>
                        select distinct * where {
                            ?property schema:domainIncludes  schema:%s .
                            ?property schema:rangeIncludes  ?exp_type .
                            ?property rdfs:label ?prop_name.
                            ?property rdfs:comment ?description
                        }""" % class_name)
    temp_dic = {}

    for row in qres:
        labels=row.labels.keys()
        labels_dic = {}
        print('Parsing %s property.' % row['prop_name'])
        for label in labels:
            labels_dic[label] = str(row[label]).replace('<a href=\"/docs/', '<a href=\"http://schema.org/docs/')
        temp_dic=__add_property(temp_dic, labels_dic)

    return temp_dic


def __get_parent_type(class_name, graph):

    print("Find parent type of %s in Schema.org" % class_name)

    qres = graph.query("""prefix schema: <http://schema.org/>
                          select ?supclass where{
                          ?class rdfs:label ?label .
                          ?class rdfs:subClassOf ?supclass .
                          filter (?label='%s')
                        }""" % class_name)
    resp_arr=[]

    for row in qres:
        resp_arr.append(str(row['supclass']))
    return resp_arr[0].replace('http://schema.org/', '')


def __get_properties(class_name, graph, properties):

    if(class_name=='Thing'):
        properties[class_name]=__get_class_props(class_name, graph)
        return properties
    else:
        temp_props = __get_class_props(class_name, graph)
        properties[class_name] = temp_props
        parent_type = __get_parent_type(class_name, graph)
        __get_properties(parent_type, graph, properties)


def get_properties_in_hierarchy(type_name):
    query_type = type_name
    g = ConjunctiveGraph()
    g.parse('http://schema.org/version/latest/schema.jsonld', format='json-ld')
    props_dic={}
    __get_properties(query_type, g, props_dic)
    return props_dic


def get_hierarchy(props_dic):
    type_hierarchy = []
    for h_type in props_dic:
        type_hierarchy.append(h_type)
    return type_hierarchy


# Function that receives an string with expected types and generates an array with each expected pype
def get_expected_type(expected_types):

    expected_types = expected_types.strip()
    expected_types = expected_types.replace('\n', '')
    expected_types = expected_types.replace(' OR ', ' ')
    expected_types = expected_types.replace(' or ', ' ')
    expected_types = expected_types.replace(',', '')
    list_of_types = expected_types.split(" ")
    i = 0
    for type in list_of_types:
        list_of_types[i] = type.strip()
        i += 1

    return list_of_types


def _parse_controlled_vocabulary(temp_cont_vocab):
    cv_parsed = {'terms':[] , 'ontologies':[]}
    element_list = temp_cont_vocab.split(',')
    for element in element_list:
        if ':' in element:
            temp_onto = element.split(":",1)
            ontology = {}
            ontology['name'] = temp_onto[0].strip()
            ontology['url'] = temp_onto[1].strip()
            cv_parsed['ontologies'].append(ontology)
        elif element != '':
            element = element.replace('LIST - ', '').strip()
            temp_term = {}
            temp_term['name'] = element
            cv_parsed['terms'].append(temp_term)
    return cv_parsed


def __get_dic_from_sheet_row(c_property):

    property_as_dic = {}

    # Set Bioschemas attributes

    property_as_dic['bsc_dec'] = c_property['BSC Description'].strip().replace('\n', ' ')
    property_as_dic['marginality'] = c_property['Marginality'].replace('\n', ' ')
    property_as_dic['cardinality'] = c_property['Cardinality'].strip().strip('\n').replace('\n', ' ')
    temp_cont_vocab = c_property['Controlled Vocabulary'].strip().replace('\n', ' ')
    property_as_dic['controlled_vocab'] = _parse_controlled_vocabulary(temp_cont_vocab)


    # Set schema.org attributes

    property_as_dic['name'] = c_property['Property'].strip().strip('\n')
    property_as_dic['expected_type'] = get_expected_type(c_property['Expected Type'])
    property_as_dic['sdo_desc'] = c_property['Description'].strip().replace('\n', ' ')
    print (property_as_dic['name'] + ':' + property_as_dic['sdo_desc'] +'\n')
    if property_as_dic['sdo_desc'] is None:
        property_as_dic['sdo_desc'] = ' ';

    return property_as_dic


def get_property_in_hierarchy(sdo_props, mapping_property):
    prop_type="new_sdo"
    for hierarchy_level in sdo_props:
        if mapping_property['name'] in sdo_props[hierarchy_level].keys():
            prop_type = hierarchy_level
            mapping_property['sdo_desc']=sdo_props[hierarchy_level][mapping_property['name']]['description']
    return {'type':prop_type, 'property': mapping_property}


def get_formatted_props(sdo_props, mapping_props, spec_name, spec_type):
    all_props= []
    bsc_props = []

    # if type only get new properties from mapping file
    if(spec_type == "Type" or spec_type == "type"):
        for mapping_property in mapping_props:
            bsc_props.append(mapping_property['name'])
            temp_prop=get_property_in_hierarchy(sdo_props, mapping_property)
            if temp_prop['type'] == "new_sdo":
                temp_prop['property']['parent'] = spec_name
            all_props.append(temp_prop['property'])
        for sdo_prop in sdo_props:
            # now get all props from schema & make them such that _layout can use them
            for sdo_prop_prop in sdo_props[sdo_prop].keys():
                if sdo_props[sdo_prop][sdo_prop_prop]['prop_name'] not in bsc_props:
                    sdo_props[sdo_prop][sdo_prop_prop]['parent'] = sdo_prop
                    sdo_props[sdo_prop][sdo_prop_prop]['name'] = sdo_props[sdo_prop][sdo_prop_prop]['prop_name']
                    # sdo_props[sdo_prop][sdo_prop_prop]['bsc_dec'] = sdo_props[sdo_prop][sdo_prop_prop]['description']
                    sdo_props[sdo_prop][sdo_prop_prop]['sdo_desc'] = sdo_props[sdo_prop][sdo_prop_prop]['description']
                    sdo_props[sdo_prop][sdo_prop_prop]['expected_type'] = sdo_props[sdo_prop][sdo_prop_prop]['exp_type']
                    all_props.append(sdo_props[sdo_prop][sdo_prop_prop])
                else:
                    for i in all_props:
                        if i['name'] == sdo_props[sdo_prop][sdo_prop_prop]['prop_name']:
                            i['parent'] = sdo_prop
        return {'properties': all_props}

    # if profile
    for mapping_property in mapping_props:
        temp_prop=get_property_in_hierarchy(sdo_props, mapping_property)
        if temp_prop['type'] == "new_sdo":
            temp_prop['property']['parent'] = spec_name
        else:
            temp_prop['property']['parent'] = temp_prop['type']
        all_props.append(temp_prop['property'])

    return {'properties': all_props}


def get_mapping_properties(mapping_sheet, spec_type):
    list_of_hashes = mapping_sheet.get_all_records(head=5)
    type_properties = []
    for c_property in list_of_hashes:
        if(c_property['Expected Type']!="" # and c_property['Description']!=""
           and c_property['Marginality']!="" and c_property['Cardinality']!=""):
            print("Parsing %s property from Google Sheets." % c_property['Property'])
            property_as_dic=__get_dic_from_sheet_row(c_property)
            type_properties.append(property_as_dic)
    return type_properties


class WorkbookParser:
    spec_metadata = {}
    bsc_specification = {}

    def __init__(self, workbook):
        if os.path.exists(workbook)
            workbook = load_workbook(workbook)
        self.workbook = workbook

    def set_gsheet_id(self, gsheet_id):
        self.gsheet_id = gsheet_id

    def set_spec_metadata(self, spec_metadata):
        self.spec_metadata = spec_metadata

    def check_url(self, spec_url):
        '''check_url doesn't exit if the address isn't found, etc.
           it just adds the string "err_404" as metadata given these cases.
        '''
        if spec_url is None: 
            return "err_404"

        response = requests.get(spec_url)
        if response.status_code == 404:
            return "err_404"
        else:
            return spec_url

    def __get_mapping_description(self, workbook=None):

        if not workbook:
            workbook = self.workbook

        # Generate values in advance
        name = self.spec_metadata['name']
        gh_base = 'https://github.com/BioSchemas/specifications/tree/master'
        use_cases_url = self.spec_metadata['use_cases_url']

        mapping_description = {}
        mapping_description['name'] = name
        print("Parsing %s Workbook" % mapping_description['name'])

        mapping_description['status'] = self.spec_metadata['status']
        mapping_description['spec_type'] = self.spec_metadata['spec_type']

        # Github Future Links
        mapping_description['gh_folder'] = '%s/%s' % (gh_base, name)
        mapping_description['gh_mapping_url'] = '%s/%s/%s_Mapping.xlsx' % (gh_base, name, name)
        mapping_description['gh_examples']= '%s/%s/examples' % (gh_base, name)
        mapping_description['gh_tasks'] = 'https://github.com/BioSchemas/bioschemas/labels/type%3A%20'+ name

        mapping_description['edit_url']='%s/%s/specification.html' % (gh_base, name)
        mapping_description['use_cases_url'] = self.check_url(use_cases_url)
        mapping_description['version'] = self.spec_metadata['version']
        
        # TODO: I don't know where to get this
        mapping_description['parent_type'] = 'CreativeWork'

        # These are from the workbook
        #TODO:WARNING this is too error prone!
        if "Specification Info" in workbook:
            mapping_sheet = workbook.get_sheet_by_name('Specification Info')
            mapping_description['subtitle'] = mapping_sheet.cell('B3').value
            mapping_description['description'] = mapping_sheet.cell('c3').value
        return mapping_description

    def get_mapping(self):

        print("Parsing %s." % self.spec_metadata['name'])

        # STOPPED HERE - this Google sheet id does not exist
        # mapping_sheet = client.open_by_key(self.gsheet_id).get_worksheet(0)

        spec_description = self.__get_mapping_description(mapping_sheet)
        sdo_props = get_properties_in_hierarchy(spec_description['parent_type'])

        spec_description['hierarchy'] = get_hierarchy(sdo_props)
        spec_description['hierarchy'].reverse()
        print_hierarchy = ' > '.join(spec_description['hierarchy'])
        print("Prepared schema.org properties for hierarchy %s" % print_hierarchy)
        print("Classifing %s properties" % spec_description['name'])
        mapping_props = get_mapping_properties(mapping_sheet,
                                               spec_description['spec_type'])


        formatted_props = get_formatted_props(sdo_props, mapping_props, spec_description['name'], spec_description['spec_type'])

        spec_description.update(formatted_props)

        return spec_description
